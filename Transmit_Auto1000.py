import sys
import os
import re
import shutil
import glob
import time
import xlwings as xw
import win32com.client
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime

def find_excel_transmittal():
    print("Searching for Transmittal Excel file...")        
    template_file = r"C:\Users\eddpa\Desktop\GoalFolder\Transmittal_TEMPLATE.xlsx" # Replace with template path used in company
    rootDirectory = os.path.abspath(".")
    rootKey = "."
    
    # Find and list all transmittal files with date in the name
    transmitt_pattern = re.compile(r"transmittal[ _-]\d{6}\.(?:xlsx)$", re.IGNORECASE)
    root_path = Path(rootKey)
    transmitt_match = []
    for p in root_path.rglob("*"):
        if p.is_file() and transmitt_pattern.search(p.name):
            transmitt_match.append(str(p.resolve()))

    # Get modification times and append to list
    transmitt_and_modtimes = []    
    if transmitt_match:
        for file_path in transmitt_match:
            mod_time = datetime.fromtimestamp(os.path.getmtime(file_path))
            transmitt_and_modtimes.append((file_path, mod_time))
            #print(f"{file_path}: {mod_time.strftime('%Y-%m-%d %H:%M:%S')}")

    # Find and copy the latest modified transmittal file to root directory
    if transmitt_and_modtimes:
        latest_transmittal = max(transmitt_and_modtimes, key=lambda x: x[1])
        latest_path = latest_transmittal[0]
        latest_filename = os.path.basename(latest_path)
        dest_path = os.path.join(rootDirectory, latest_filename)        
        print(f"Transmittal files found... \nThe latest transmittal excel is: {latest_path}")

        # Copy file only if source and destination are different
        if os.path.abspath(latest_path) != os.path.abspath(dest_path):
            shutil.copy(latest_path, rootDirectory)
            print("Latest Transmittal Excel file moved into root directory.")
            return dest_path
        else:
            # File is already in the root directory
            print("Latest Transmittal file is already in the root directory.")
            return latest_path
    else:
        print("Transmittal Excel file not found. Copying from Template source...")
    
    # Copy template file if no transmittal found
    try:
        shutil.copy(template_file, rootDirectory)
        print(f"Template Transmittal'{template_file}' copied successfully to '{rootDirectory}'")
    except FileNotFoundError:
        print("Error: The source file was not found.")
    except PermissionError:
        print("Error: You do not have permission to write to the destination.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
    return os.path.join(rootDirectory, "Transmittal_TEMPLATE.xlsx")

def Catch_Drawings():
    print("Searching for drawings in current folder and subfolders...")    
    # Define pattern of file name that ends in .pdf and has a character or digit inside square brackets
    glob_pattern = "*[[]?*[]]*.pdf"

    # Strict regex to validate bracket content
    regex_pattern = re.compile(r"\[[A-Za-z0-9]+\].*\.pdf$", re.IGNORECASE)

    # Find all PDFs matching the glob pattern
    all_matches = glob.glob(glob_pattern)
    
    # Filter with regex to ensure bracket content is alphanumeric only
    rawListDrawings = []
    for file_path in all_matches:
        filename = os.path.basename(file_path)
        if regex_pattern.search(filename):
            # Store full absolute path
            rawListDrawings.append(os.path.abspath(file_path))

    if rawListDrawings:
        print(f"Found {len(rawListDrawings)} drawings in the current folder.") 
        return rawListDrawings
    else:        
        raise ValueError("No PDF drawings found to process.")

def Request_Get_Date():
    rootDirectory = os.path.abspath(".")
    print("Choose a date for the transmittal (DD/MM/YY), from the following options: ")
    while True:
        # Print the menu options        
        print("1. Today's Date")
        print("2. Enter a Custom Date")
        print("3. Exit Program")

        # Prompt for user input
        choice = input("Enter your choice (1-3): ")
        # Activate choice
        if choice == '1':
            from datetime import datetime
            now = datetime.now()
            date_string = now.strftime("%d/%m/%y")
            break
        elif choice == '2':
            date_string = input("Enter the date in the following format DD/MM/YY : ")
            if len(date_string) != 8 or date_string[2] != '/' or date_string[5] != '/':
                print("Invalid date format. Please try again.")
                input("Press Enter to continue...")
                continue
            break
        elif choice == '3':
            print("Exiting the program. Goodbye!")
            sys.exit(0)
        else:
            print("Invalid choice. Please enter a number between 1 and 3.")
            input("Press Enter to continue...")
    # Load transmittal Excel file
    transmittal = find_excel_transmittal()
    workbook = load_workbook(transmittal)

    # Check if 'CIVIL' sheet exists
    if 'CIVIL' in workbook.sheetnames:
        print("CIVIL sheet found.")
        worksheet = workbook['CIVIL']                       
    else:
        raise ValueError("Sheet 'CIVIL' not found in the Excel file.")

    # Get the date from user and overwrite empty date cells    
    dateStrings = date_string.split('/')
    dateParts_int = [int(s) for s in dateStrings]    
    dateRow_index = 1    
    for date_cells in worksheet.iter_rows(min_row=dateRow_index, max_row=dateRow_index, min_col=5, max_col=30):    
        for cell in date_cells:
            # Check if date already exists and use that column for updating revisions
            if cell.value == dateParts_int[0] and worksheet.cell(row=cell.row + 1, column=cell.column).value == dateParts_int[1] and worksheet.cell(row=cell.row + 2, column=cell.column).value == dateParts_int[2]:
                print(f"Date {date_string} already exists in the transmittal at cell {cell.coordinate}.")                                
                print(f"Column: {cell.column} will be used for revisions update.")
                break            
            elif cell.value is None:
                print("New date cell is: " + cell.coordinate)
                for i, part in enumerate(dateParts_int):
                    target_cell = dateRow_index + i
                    worksheet.cell(row=target_cell, column=cell.column, value=part)                    
                break

    # Save the modified workbook
    dateParts = [s for s in dateStrings]
    fileName_date = f"{dateParts[2]}{dateParts[1]}{dateParts[0]}"
    output_filename = f"Transmittal {fileName_date}.xlsx"
    workbook.save(output_filename)
    print(f"Transmittal excel file saved as '{output_filename}'.")

    return os.path.join(rootDirectory, output_filename)

def Update_Transmittal():
    # Load transmittal Excel file
    print("Loading new Transmittal Excel file...")
    transmittal = Request_Get_Date()
    # transmittal = r"C:\Users\eddpa\Desktop\Transmittal_Auto1000\Transmittal 250418.xlsx"    
    workbook = xw.Book(transmittal)

    print("Choose sheet to process:")
    while True:
        # Print the menu options
        print("1. CIVIL")
        print("2. STRUCTURE")
        print("3. ARCHITECT")
        print("4. Exit Program")

        # Prompt for user input
        choice = input("Enter your choice (1-3): ")
        # Activate choice
        if choice == '1':
            # Check if 'CIVIL' sheet exists. Select it if it does
            if workbook.sheets["CIVIL"]:
                print("CIVIL sheet found.")
                worksheet = workbook.sheets["CIVIL"]
                sheet_name = "CIVIL"
            else:
                raise ValueError("Sheet 'CIVIL' not found in the Excel file. Check sheet name spelling")
            break
        elif choice == '2':
            # Check if 'ARCHITECT' sheet exists. Select it if it does
            if workbook.sheets["ARCHITECT"]:
                print("ARCHITECT sheet found.")
                worksheet = workbook.sheets["ARCHITECT"]
                sheet_name = "ARCHITECT"
            else:
                raise ValueError("Sheet 'ARCHITECT' not found in the Excel file. Check sheet name spelling")
            break
        elif choice == '3':
            # Check if 'STRUCTURE' sheet exists. Select it if it does
            if workbook.sheets["STRUCTURE"]:
                print("STRUCTURE sheet found.")
                worksheet = workbook.sheets["STRUCTURE"]
                sheet_name = "STRUCTURE"
            else:
                raise ValueError("Sheet 'ARCHITECT' not found in the Excel file. Check sheet name spelling")
            break
        elif choice == '4':
            print("Exiting the program. Goodbye!")
            sys.exit(0)
        else:
            print("Invalid choice. Please enter a number between 1 and 3.")
            input("Press Enter to continue...")


    # Get revision column index from Excel by last date reference
    rev_column = None
    dateRow_index = 1
    for col_index in range(5, 31):
        cell = worksheet.range((dateRow_index, col_index))  # (row, col) tuple
        if cell.value is None:
            rev_column = cell.column - 1  # Revision column is the one before the empty cell
            print(f"Found revision column at index: {rev_column}")
            break # If found, Exit inner loop
    
    if rev_column is None:        
        raise ValueError("Revision column not found in worksheet.")

    ######### Compare PDF vs Excel names, update revision when matched and add new drawings ##########
    # Get revision from raw list of PDF drawings
    rawlist_PDF = Catch_Drawings()

    for file_Name in rawlist_PDF:
        print(f"\nProcessing drawing file: {file_Name}")
        # Define regex patterns to extract project number, revision, and drawing name
        pjtNo_pattern = r"\\([^\\]*)-[A-Z]" # Pattern to catch anything before department letter (C, A or S) with hyphen)
        dwg_pattern = r"\d{5}-(.*) \[" # Pattern to catch anything between 5 digits with a hyphen and " ["
        rev_pattern = r"\[(.*)\]" # Pattern to catch anything between "[ " and "]"        
        name_pattern = r"\] (.*)\.pdf" # Pattern to catch anything between "] " and ".pdf"
        pjtNo_match = re.search(pjtNo_pattern, file_Name)
        dwg_match = re.search(dwg_pattern, file_Name)
        rev_match = re.search(rev_pattern, file_Name)
        name_match = re.search(name_pattern, file_Name)
        
        # Skip file if pattern not found
        if not (pjtNo_match and dwg_match and rev_match and name_match):
            print(f"Could not parse all details from '{file_Name}'. Skipping.")
            continue # Skip to the next file
        
        # Strip found values from " " and store them
        project_No = pjtNo_match.group(1).strip() 
        drawing_No = dwg_match.group(1).strip() 
        revision = rev_match.group(1).strip() 
        drawing_Name = name_match.group(1).strip()

        # Get drawing group number and count number (e.g. drawing C-05-11, has group 5 and count 11)
        drawing_No_parts = drawing_No.split('-')
        drawing_Group = int(drawing_No_parts[1]) # e.g., 5
        drawing_Count = int(drawing_No_parts[2]) # e.g., 11
        
        # Compare drawing names from list_PDF and drawing names from cells in Excel
        excel_Match = False # Flag to track if the drawing name matched in Excel           
        for row_index in range(24, 151):
            name_cell = worksheet.range((row_index, 3))  # (row, col) tuple
            # If there is a match, update revision in the same row at the revision column
            if name_cell.value and str(name_cell.value).strip() == drawing_Name:
                rev_cell = worksheet.range((name_cell.row, rev_column))  # Revision is updated at this cell coordinate
                rev_cell.value = revision
                print(f"Matched '{drawing_Name}'. Updated revision to '{revision}' at row {name_cell.row}.")
                excel_Match = True
                break  # Exit the loop for this drawing file because a match was found

        # If not a match, add new drawing name into new row
        match_counter = 0
        if not excel_Match :
            print(f"No match found for drawing: {drawing_Name}. Adding new entry...")
            # Find drawing number in column B to determine where to add new drawing
            for row_index in range(24, 151):
                # Get drawing group number and count number from cell value (e.g. drawing C-05-11, has group 5 and count 11)
                drawing_No_cell = worksheet.range((row_index, 2)).value
                if not drawing_No_cell:
                    continue
                parts = str(drawing_No_cell).split('-')
                # defensive: ensure expected format
                if len(parts) < 3:
                    continue
                try:
                    drawing_Group_cell = int(parts[1])
                except ValueError:
                    continue

                if drawing_Group == drawing_Group_cell:
                    match_counter += 1
                    last_match_row = row_index  # update each time a match is found
            
            if match_counter > 0:
                # There are existing drawings in the same group, find correct position to insert
                for row_index in range(24, 151):
                    drawing_No_cell = worksheet.range((row_index, 2)).value
                    # CRITICAL: Skip empty cells to avoid NoneType error
                    if not drawing_No_cell:
                        continue                    
                    drawing_No_parts_cell = drawing_No_cell.split('-')                    
                    # Ensure the split produced at least 3 parts
                    if len(drawing_No_parts_cell) < 3:
                        continue                    
                    try:
                        drawing_Group_cell = int(drawing_No_parts_cell[1])  # e.g., 5
                        drawing_Count_cell = int(drawing_No_parts_cell[2])  # e.g., 11
                    except ValueError:
                        continue                                        
                    
                    # Check where to insert new drawing if drawing_Count is less than existing ones
                    if drawing_Group == drawing_Group_cell and drawing_Count < drawing_Count_cell:                        
                        worksheet.api.Rows(row_index).Insert()                                  # insert a blank row
                        worksheet.api.Rows(row_index + 1).Copy(worksheet.api.Rows(row_index))   # copy formulas/format from row below
                        
                        worksheet.range((row_index, 1)).value = project_No  # Add project number                        
                        worksheet.range((row_index, 2)).value = drawing_No  # Add drawing number                        
                        worksheet.range((row_index, 3)).value = drawing_Name  # Add drawing name                        
                        worksheet.range((row_index, rev_column)).value = revision  # Set revision from PDF
                        print(f"Added new drawing {drawing_Name} at row {row_index} with revision {revision}.")                            
                        break  # Exit the loop after adding the new drawing to avoid multiple additions
                    elif row_index == last_match_row:
                        # If reached last match row and no smaller count found, add after last match
                        next_row = last_match_row + 1                        
                        worksheet.api.Rows(next_row).Insert()                                 # insert a blank row
                        worksheet.api.Rows(next_row + 1).Copy(worksheet.api.Rows(next_row))   # copy formulas/format from row below
                        
                        worksheet.range((next_row, 1)).value = project_No  # Add project number                        
                        worksheet.range((next_row, 2)).value = drawing_No  # Add drawing number                        
                        worksheet.range((next_row, 3)).value = drawing_Name  # Add drawing name                        
                        worksheet.range((next_row, rev_column)).value = revision  # Set revision from PDF
                        print(f"Added new drawing {drawing_Name} at row {next_row} with revision {revision}.")                            
                        break  # Exit the loop after adding the new drawing to avoid multiple additions
            else:
                # No existing drawings in the same group, add to the first empty row found
                for empty_row in range(24, 151):
                    if worksheet.range((empty_row, 2)).value is None:
                        worksheet.range((empty_row, 1)).value = project_No  # Add project number
                        worksheet.range((empty_row, 2)).value = drawing_No  # Add drawing number
                        worksheet.range((empty_row, 3)).value = drawing_Name  # Add drawing name
                        worksheet.range((empty_row, rev_column)).value = revision  # Set initial revision to 1
                        print(f"Added new drawing {drawing_Name} at row {empty_row} with revision {revision}.")                            
                        break  # Exit the loop after adding the new drawing to avoid multiple additions

    # Extract transmittal file name from path
    pattern = r"transmittal \d{6}\.xlsx"
    match = re.search(pattern, transmittal, re.IGNORECASE)
    if match:
        m_filename = match.group()
    else:
        print("Transmittal filename does not match expected pattern.")
        print("Please check the name matches: Transmittal YYMMDD.xlsx")
        workbook.close()
        return None, None
    
    transmittal_filename = Path(m_filename)
    transmittal_filename.parent.mkdir(parents=True, exist_ok=True)   # ensure folder exists

    # Save the modified workbook    
    workbook.save(transmittal_filename)
    print(f"\nChanges saved in Transmittal excel '{transmittal_filename}'.")

    # CRITICAL: Close the workbook before returning so it's not locked
    try:
        workbook.close()
        print("Workbook closed successfully.")
    except Exception as e:
        print(f"Error closing workbook: {e}")
    
    # Small delay to allow Excel to fully terminate        
    time.sleep(5)

    return transmittal, sheet_name

def Save_as_PDF():
    """
    Saves a specific worksheet from an Excel file to a PDF with A4 format
    by controlling the Excel application via COM.
    """
    excel_path, sheet_name = Update_Transmittal()

    # Check if Overwrite_Transmittal() succeeded
    if excel_path is None or sheet_name is None:
        print("Error: Could not generate transmittal file. Aborting PDF export.")
        return

    currentDir = os.path.abspath(".")
    pattern = r"\btransmittal (\d{6})\.xlsx\b"
    match = re.search(pattern, excel_path, re.IGNORECASE)
    if match:
        pdf_name = f"Transmittal {match.group(1)}.pdf"
        pdf_path = os.path.join(currentDir, pdf_name)
    else:
        print("Transmittal filename does not match expected pattern.")
        print("Please check the name matches: Transmittal YYMMDD.xlsx")
        return

    print(f"\nConverting '{sheet_name}' from '{excel_path}' to PDF using MS Excel...")

    excel = None  # Initialize excel variable
    workbook = None # Initialize workbook variable
    try:
        # Get absolute paths, which are required for COM
        excel_abs_path = os.path.abspath(excel_path).replace("/", "\\")
        pdf_abs_path = os.path.abspath(pdf_path).replace("/", "\\")

        # CRITICAL: Verify the file exists before attempting to open
        if not os.path.exists(excel_abs_path):
            print(f"Error: File not found at '{excel_abs_path}'")
            return

        print(f"Opening file: {excel_abs_path}")

        # Start an instance of Excel
        excel = win32com.client.Dispatch("Excel.Application")
        # Keep the application hidden
        excel.Visible = False
        excel.DisplayAlerts = False  # Suppress warning dialogs

        # Open the workbook with error handling
        try:
            workbook = excel.Workbooks.Open(excel_abs_path, ReadOnly=True, IgnoreReadOnlyRecommended=True)
        except Exception as e:
            print(f"Error opening workbook: {e}")
            print(f"File path: {excel_abs_path}")
            raise

        # Select the specific worksheet
        worksheet = workbook.Worksheets[sheet_name]

        # --- Set Page Setup ---
        # xlPaperA4 has a value of 9
        worksheet.PageSetup.PaperSize = 9 
        # Default margins are used automatically.

        # --- Export to PDF ---
        # xlTypePDF has a value of 0
        worksheet.ExportAsFixedFormat(0, pdf_abs_path)
        
        print(f"Successfully saved PDF to '{pdf_abs_path}'")

        workbook.Close(SaveChanges=False)
        excel.Quit()
        print("Excel application closed.")

    except Exception as e:
        print(f"An error occurred during PDF conversion: {e}")        

if __name__ == "__main__": 
    print("Transmit_Auto1000 Start")    
    Save_as_PDF()
    print("Created by Edd Palencia-Vanegas - June 2024. All rights reserved.")
    print("Version 5.1 - 13/01/2026")